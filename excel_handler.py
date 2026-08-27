"""
excel_handler.py — Safe Excel updates via win32com

Handles reading and writing to gieco_fleet_master_2026.xlsx:
- Reads master vehicle list with chassis and plate numbers
- Matches OCR extracted data against sheet rows
- Prepares and validates changes (diff generation)
- Applies updates using win32com (Excel COM automation) to preserve all formulas
- Highlights updated rows/cells in Soft Green
- Creates automatic timestamped backups before applying changes

Usage:
    handler = ExcelHandler("gieco_fleet_master_2026.xlsx")
    match = handler.find_vehicle("34271")
    diff = handler.prepare_update(match, ocr_data)
    handler.apply_update(diff, dry_run=True)
"""

import os
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
import shutil

from datetime import datetime, date
from dataclasses import dataclass, field
from typing import Optional

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment

# Soft audit green fill for modified cells and rows
AUDIT_HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
AUDIT_ROW_KEY_FILL = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")



# Column indices in the master sheet (1-indexed for openpyxl)
# These map to the 35 columns in سيارات 2026
COL = {
    "تاريخ الاستلام (receipt)":    1,   # Col A
    "مكتب تامينات السيارة":        2,   # Col B
    "مؤمن عليه فى مكتب":           3,   # Col C
    "كود الموظف":                   4,   # Col D
    "رقم تأميني":                   5,   # Col E
    "الاسم":                        6,   # Col F
    "الرقم القومي":                 7,   # Col G
    "على سيارة":                    8,   # Col H
    "المسمى الوظيفي":               9,   # Col I
    "تاريخ الالتحاق":              10,   # Col J
    "تاريخ انتهاء الرخصة":         11,   # Col K
    "قبض 12":                      12,   # Col L
    "ليبيا":                       13,   # Col M
    "ساركي":                       14,   # Col N
    "بدلا من السائق":              15,   # Col O
    "تاريخ الاستقالة":             16,   # Col P
    "سبب التغيير":                 17,   # Col Q
    "رقم السيارة":                 18,   # Col R (Plate Number)
    "الشاسيه":                     19,   # Col S (Chassis Number)
    "نهاية رسوم الضريبة":          20,   # Col T
    "حالة السيارة":                21,   # Col U
    "نوع السيارة":                 22,   # Col V
    "كود شهادة":                   23,   # Col W  ** TARGET **
    "رقم السيارة القديم":          24,   # Col X
    "نهاية التصريح/ الترخيص":      25,   # Col Y  ** TARGET **
    "تاريخ ارسالها التأمينات":     26,   # Col Z
    "تاريخ الاستلام":              27,   # Col AA ** TARGET **
    "الموقف":                      28,   # Col AB ** TARGET **
    "تاريخ اصدار الشهادة":         29,   # Col AC ** TARGET **
    "من (validity start)":         30,   # Col AD ** TARGET **
    "الى (validity end)":          31,   # Col AE ** TARGET **
    "المدة بالشهور":               32,   # Col AF (FORMULA — DO NOT TOUCH)
    "ملاحظات":                     33,   # Col AG
    "مكتب (duplicate)":            34,   # Col AH
    "كود الموظف (duplicate)":      35,   # Col AI
}

# The target columns we write to (1-indexed)
TARGET_COLS = {
    "تاريخ الاستلام":      27,
    "من":                  30,
    "الى":                 31,
    "الرقم التأميني":      36,  # Col AJ
}

MASTER_SHEET = "سيارات 2026"
HEADER_ROWS = 2  # Rows 1-2 are headers, data starts at row 3


@dataclass
class VehicleMatch:
    """Represents a matched vehicle record in the Excel sheet."""
    row_number: int
    chassis: str
    plate: str
    driver_name: str
    employee_code: str
    office: str
    vehicle_type: str
    current_cert_code: Optional[int]
    current_expiry: Optional[str]
    current_date_to: Optional[datetime]
    match_type: str  # 'chassis' or 'plate'


@dataclass
class UpdateDiff:
    """Represents the planned changes before applying to Excel."""
    row_number: int
    driver_name: str
    office: str
    changes: dict = field(default_factory=dict)
    # changes maps: col_name -> {"old": old_val, "new": new_val, "col": col_number}


class ExcelHandler:
    """
    Read/Write manager for the GIECO vehicle insurance Excel workbook.

    On initialization, loads the workbook and caches the master sheet data.
    """

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Load with data_only=False to preserve formulas on write, keep_links=False to avoid external link bloat/corruption
        self.wb = openpyxl.load_workbook(excel_path, data_only=False, keep_links=False)
        self.ws = self.wb[MASTER_SHEET]
        self.max_row = self.ws.max_row

        # Also load evaluated values (data_only=True) to read calculated formulas like cert_code (Col W) and driver names
        try:
            self.wb_data = openpyxl.load_workbook(excel_path, data_only=True, keep_links=False)
            self.ws_data = self.wb_data[MASTER_SHEET]
        except Exception as e:
            print(f"[WARN] Failed to load data_only workbook: {e}")
            self.wb_data = None
            self.ws_data = self.ws


        print(f"[ExcelHandler] Loaded: {excel_path}")
        print(f"[ExcelHandler] Master sheet: {MASTER_SHEET}, rows: {self.max_row}")


    def find_vehicle(self, chassis_or_plate: str) -> list[VehicleMatch]:
        """
        Search the master sheet for a vehicle by Chassis Number or Plate Digits.

        Matching priority:
            1. Exact chassis match  (Col 19) — safest, highest confidence.
            2. Partial chassis match (Col 19) — only if no exact match found;
               flagged as 'chassis (partial - review required)'.
            3. Exact plate-digit match (Col 18) — only if both chassis passes fail.

        Args:
            chassis_or_plate: Chassis number string or plate digit string.

        Returns:
            List of VehicleMatch objects (may be 0, 1, or 2 for dual-driver vehicles).
        """
        matches: list = []
        search = str(chassis_or_plate).strip()

        # Pass 1: Exact chassis match (Col 19)
        for row in range(HEADER_ROWS + 1, self.max_row + 1):
            val = self.ws.cell(row=row, column=19).value
            if val is not None:
                clean_val = str(val).strip()
                if search == clean_val:
                    match = self._build_match(row, "chassis")
                    if match:
                        matches.append(match)

        # Pass 2: Partial chassis match — only if exact pass found nothing
        if not matches:
            for row in range(HEADER_ROWS + 1, self.max_row + 1):
                val = self.ws.cell(row=row, column=19).value
                if val is not None:
                    clean_val = str(val).strip()
                    if clean_val and search != clean_val and (
                        search in clean_val or clean_val.endswith(search)
                    ):
                        match = self._build_match(row, "chassis (partial - review required)")
                        if match:
                            matches.append(match)

        # Pass 3: Exact plate-digit match (Col 18) — fallback only
        if not matches:
            for row in range(HEADER_ROWS + 1, self.max_row + 1):
                val = self.ws.cell(row=row, column=18).value
                if val is not None and str(val).strip():
                    digits = "".join(c for c in str(val) if c.isdigit())
                    search_digits = "".join(c for c in search if c.isdigit())
                    if search_digits and digits == search_digits:
                        match = self._build_match(row, "plate")
                        if match:
                            matches.append(match)

        return matches

    @staticmethod
    def _safe_int(val) -> Optional[int]:
        """Safely convert a cell value to int, handling formulas and None."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip()
        if s.startswith("=") or not s or s == "#N/A":
            return None  # Formula or empty
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None

    def _build_match(self, row: int, match_type: str) -> Optional[VehicleMatch]:
        """Build a VehicleMatch from a given row."""
        chassis_val = self.ws.cell(row=row, column=19).value
        plate_val = self.ws.cell(row=row, column=18).value

        # Read display values from evaluated data workbook if available
        name_val = self.ws_data.cell(row=row, column=6).value or self.ws.cell(row=row, column=6).value
        emp_code = self.ws_data.cell(row=row, column=4).value or self.ws.cell(row=row, column=4).value
        office_val = self.ws_data.cell(row=row, column=2).value or self.ws.cell(row=row, column=2).value
        vtype_val = self.ws_data.cell(row=row, column=22).value or self.ws.cell(row=row, column=22).value
        cert_code = self.ws_data.cell(row=row, column=23).value or self.ws.cell(row=row, column=23).value
        expiry_val = self.ws_data.cell(row=row, column=25).value or self.ws.cell(row=row, column=25).value
        date_to_val = self.ws_data.cell(row=row, column=31).value or self.ws.cell(row=row, column=31).value

        # Skip header-like rows
        if str(chassis_val).strip() in ("الشاسيه", "None", ""):
            return None

        # Parse date_to_val to datetime if it's a string, or keep it if it's already datetime
        dt_to = None
        if isinstance(date_to_val, datetime):
            dt_to = date_to_val
        elif isinstance(date_to_val, str) and date_to_val.strip() and not date_to_val.startswith("="):
            try:
                dt_to = datetime.strptime(date_to_val.strip()[:10], "%Y-%m-%d")
            except ValueError:
                pass

        return VehicleMatch(
            row_number=row,
            chassis=str(chassis_val or "").strip(),
            plate=str(plate_val or "").strip(),
            driver_name=str(name_val or "").strip(),
            employee_code=str(emp_code or "").strip(),
            office=str(office_val or "").strip(),
            vehicle_type=str(vtype_val or "").strip(),
            current_cert_code=self._safe_int(cert_code) if self._safe_int(cert_code) is not None else (str(cert_code).strip() if cert_code and not str(cert_code).startswith("=") else None),
            current_expiry=str(expiry_val) if expiry_val and not str(expiry_val).startswith("=") else None,
            current_date_to=dt_to,
            match_type=match_type,
        )

    def prepare_update(self, matches: list[VehicleMatch],
                       ocr_data: dict) -> list[UpdateDiff]:
        """
        Prepare the exact changes to apply to Excel rows.

        Args:
            matches: List of VehicleMatch (1 or 2 rows for same chassis).
            ocr_data: Dict from ocr_engine.extract_certificate_data().

        Returns:
            List of UpdateDiff objects showing old vs new values.
        """
        today = date.today()
        diffs = []
        for match in matches:
            diff = UpdateDiff(
                row_number=match.row_number,
                driver_name=match.driver_name,
                office=match.office,
            )

            # Col 27: تاريخ الاستلام = from UI
            old_receipt = self.ws.cell(row=match.row_number, column=27).value
            diff.changes["تاريخ الاستلام"] = {
                "old": str(old_receipt) if old_receipt else None,
                "new": ocr_data.get("receipt_date") or today.isoformat(),
                "col": 27,
            }

            # Col 30: من (validity start)
            old_from = self.ws.cell(row=match.row_number, column=30).value
            diff.changes["من"] = {
                "old": str(old_from) if old_from else None,
                "new": ocr_data.get("date_from"),
                "col": 30,
            }

            # Col 31: الى (validity end)
            old_to = self.ws.cell(row=match.row_number, column=31).value
            diff.changes["الى"] = {
                "old": str(old_to) if old_to else None,
                "new": ocr_data.get("date_to"),
                "col": 31,
            }
            
            # Col 36: الرقم التأميني للسيارة (Insurance Number) - Col AJ
            old_ins = self.ws.cell(row=match.row_number, column=36).value
            diff.changes["الرقم التأميني"] = {
                "old": str(old_ins) if old_ins else None,
                "new": ocr_data.get("insurance_no"),
                "col": 36,
            }

            diffs.append(diff)

        return diffs

    def apply_update(self, diffs: list[UpdateDiff],
                     output_path: Optional[str] = None,
                     dry_run: bool = False) -> str:
        """
        Apply the prepared changes to the Excel workbook using native Excel via win32com
        to guarantee 100% preservation of all array formulas, macros, and external links,
        and highlight modified cells in Yellow.
        """
        print("\n" + "=" * 70)
        print("EXCEL UPDATE REPORT")
        print("=" * 70)

        for diff in diffs:
            print(f"\n  Row {diff.row_number} | {diff.driver_name} | {diff.office}")
            print(f"  {'─' * 60}")

            for col_name, change in diff.changes.items():
                old = change["old"]
                new = change["new"]
                col = change["col"]
                marker = "→" if old != new else "="
                print(f"    Col {col:2d} ({col_name:20s}): {old} {marker} {new}")

        if dry_run:
            print("\n  [DRY RUN] No changes written to file.")
            return "DRY_RUN"

        # Save to output path
        if output_path is None:
            base, ext = os.path.splitext(self.excel_path)
            output_path = f"{base}_updated{ext}"

        # Create a fresh timestamped backup before every write run
        from datetime import datetime as _dt
        base, ext = os.path.splitext(self.excel_path)
        backup_path = f"{base}_backup_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        shutil.copy2(self.excel_path, backup_path)
        print(f"\n  [BACKUP] Created: {backup_path}")

        import win32com.client
        import pythoncom

        # We must use absolute paths for win32com
        abs_in = os.path.abspath(self.excel_path)
        abs_out = os.path.abspath(output_path)

        # Copy original to output path so we edit in place
        if abs_in != abs_out:
            shutil.copy2(abs_in, abs_out)

        # Guard COM objects so finally clause is safe even if init fails
        excel = None
        wb = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(abs_out)
            ws = wb.Sheets(MASTER_SHEET)

            for diff in diffs:
                # Highlight the entire updated row in Bright Yellow
                row_range = ws.Range(ws.Cells(diff.row_number, 1), ws.Cells(diff.row_number, 36))
                row_range.Interior.Color = 65535  # RGB(255,255,0) => 0x00FFFF => 65535

                for col_name, change in diff.changes.items():
                    new_val = change["new"]
                    col_num = change["col"]

                    # Convert ISO date strings to MM/DD/YYYY for native Excel
                    if new_val and isinstance(new_val, str) and len(new_val) == 10:
                        try:
                            parts = new_val.split("-")
                            if len(parts) == 3:
                                new_val = f"{parts[1]}/{parts[2]}/{parts[0]}"
                        except (ValueError, IndexError):
                            pass

                    target_cell = ws.Cells(diff.row_number, col_num)
                    target_cell.Value = new_val
                    target_cell.Interior.Color = 65535

            wb.Save()

        finally:
            # Guard each step: only call if successfully initialized
            if wb is not None:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

        print(f"\n  [SAVED] Updated workbook via Excel Interop: {output_path}")
        return output_path

    def print_vehicle_info(self, match: VehicleMatch):
        """Print a formatted summary of a matched vehicle record."""
        print(f"\n  {'─' * 50}")
        print(f"  VEHICLE FOUND (Row {match.row_number})")
        print(f"  {'─' * 50}")
        print(f"    Chassis:       {match.chassis}")
        print(f"    Plate:         {match.plate}")
        print(f"    Driver:        {match.driver_name}")
        print(f"    Employee Code: {match.employee_code}")
        print(f"    Office:        {match.office}")
        print(f"    Vehicle Type:  {match.vehicle_type}")
        print(f"    Cert Code:     {match.current_cert_code}")
        print(f"    Current Expiry:{match.current_expiry}")
        print(f"    Match Type:    {match.match_type}")


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_handler.py <excel_path> [chassis_search]")
        sys.exit(1)

    handler = ExcelHandler(sys.argv[1])

    if len(sys.argv) >= 3:
        search = sys.argv[2]
        print(f"\nSearching for: {search}")
        matches = handler.find_vehicle(search)

        if matches:
            print(f"\nFound {len(matches)} match(es):")
            for m in matches:
                handler.print_vehicle_info(m)
        else:
            print(f"\n  ❌ No vehicle found for: {search}")
